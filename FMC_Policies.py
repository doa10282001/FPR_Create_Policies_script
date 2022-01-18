import requests
import json
from requests.auth import HTTPBasicAuth
import openpyxl
import time

requests.packages.urllib3.disable_warnings()

def fmc_gen_token(username,passwd):
    global response, FMC_token, Domain_UUID, FMC_refresh_token
    token_api_path = "/api/fmc_platform/v1/auth/generatetoken"
    url = Server_url + token_api_path
    conn = r.post(url, auth=HTTPBasicAuth(username, password), verify=False)
    FMC_token = conn.headers["X-auth-access-token"]
    Domain_UUID = conn.headers["DOMAIN_UUID"]
    FMC_refresh_token = conn.headers["X-auth-refresh-token"]
    return FMC_token, Domain_UUID, FMC_refresh_token

def get_url_data(api_path):
    url = Server_url + api_path
    data = r.get(url, headers=headers, verify=False)
    json_data = json.loads(data.text)
    return json_data

def post_url_data(name, api_path, post_data):
    url = Server_url + api_path
    time.sleep(1)
    data = r.post(url, data=json.dumps(post_data), headers=headers, verify=False)
    if ((data.status_code == 200) or (data.status_code == 201)):
        print('post ' + name + ' Success, Fin')
    else:
        json_data = json.loads(data.text)
        try:
            if 'already exists' in json_data['error']['messages'][0]['description']:
                print('failed to Post , already exists')
        except:
            print(data.text)

def get_version():
    version_api = "/api/fmc_platform/v1/info/serverversion"
    version_json = get_url_data(version_api)
    FMC_Version = version_json['items'][0]['serverVersion']
    return FMC_Version

def get_current_Policy_UUID():
    get_Policy_UUID_url = '/api/fmc_config/v1/domain/' + Domain_UUID + '/policy/accesspolicies'
    Policy_UUID_json = get_url_data(get_Policy_UUID_url)
    for i in Policy_UUID_json['items']:
        print('name :', i['name'])
        ID = i['id']
        print('ID :', ID)
        Policy_UUID.append(ID)

def get_Network_object_ID(net_name) -> dict:
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/object/networkaddresses'
    get_ID_json = get_url_data(api_path)
    for i in get_ID_json['items']:
        if net_name == i['name']:
            return { "type": i['type'],
                     "name": i['name'],
                     "id": i['id']}

def get_Network_object_Group_ID(net_name) ->dict:
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/object/networkgroups'
    get_ID_json = get_url_data(api_path)
    for i in get_ID_json['items']:
        if net_name == i['name']:
            return { "type": i['type'],
                     "name": i['name'],
                     "id": i['id']}

def get_zone_UUID(zone_name):
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/object/securityzones'
    get_zone_json = get_url_data(api_path)
    for i in get_zone_json['items']:
        if zone_name == i['name']:
            return i['id']

def get_url_object_ID(url_name):
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/object/urls'
    get_ID_json = get_url_data(api_path)
    for i in get_ID_json['items']:
        if url_name == i['name']:
            return { "type": i['type'],
                     "id": i['id'],
                     "name":i['name']}

def get_urlGroup_object_ID(url_name):
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUD + '/object/urlgroups'
    get_ID_json = get_url_data(api_path)
    for i in get_ID_json['items']:
        if url_name == i['name']:
            return { "type": i['type'],
                     "id": i['id'],
                     "name":i['name']}

def port_object(port_name):
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/object/protocolportobjects'
    port_info_json = get_url_data(api_path)
    for i in port_info_json['items']:
        if port_name == i['name']:
            protocol_url = i['links']['self']
            port_info_d = r.get(protocol_url, headers = headers, verify = False)
            json_port_info_d = json.loads(port_info_d.text)
            return {
                'type':json_port_info_d['type'],
                'protocol':json_port_info_d['protocol'],
                'name':json_port_info_d['name'],
                'id':json_port_info_d['id']
            }

def research_port_object(port_name):
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/object/protocolportobjects'
    port_info_json = get_url_data(api_path)
    for i in port_info_json['paging']:
        if 'next' in i:
            next_port_info_json = port_info_json['paging']['next']
    for i in next_port_info_json:
        port_info_d = r.get(i, headers = headers, verify = False)
        json_port_info_d = json.loads(port_info_d.text)
        for i in json_port_info_d['items']:
            if port_name == i['name']:
                return {
                    'type':i['type'],
                    'id':i['id']
            }

def port_object_Group_ID(port_name):
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/object/portobjectgroups'
    port_info_json = get_url_data(api_path)
    for i in port_info_json['items']:
        if port_name == i['name']:
            return{'name':i['name'],
                   'type':i['type'],
                   'id':i['id']}

def get_ICMP_UUID():
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/object/icmpv4objects'
    ICMP_UUID_json = get_url_data(api_path)
    for i in ICMP_UUID_json['items']:
        if i['name'] == "icmpv4_obj1":
            return { "type": i['type'],
                     "id" : i['id']}

def create_network_object(name,ip_address, Description = None):
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/object/networks'
    post_data = {
        "name": name,
        "value": ip_address,
        "overridable": False,
        "description": Description,
        "type": "Network"
        }
    post_url_data(name, api_path, post_data)
    #json_Create_Network_object = json.loads(Create_Network_object.text)

def create_network_object(name,ip_address, Description = None):
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/object/networks'
    post_data = {
        "name": name,
        "value": ip_address,
        "overridable": False,
        "type": "Network"
        }
    if Description != "None":
        post_data["description"] = Description
    post_url_data(name, api_path, post_data)
    #json_Create_Network_object = json.loads(Create_Network_object.text)

def create_Networkgroups_object(name, post_data):
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/object/networkgroups'
    post_url_data(name, api_path, post_data)

def create_port_object(name, protocol, port, Description = None):
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/object/protocolportobjects'
    post_data = {
        "name": name,
        "protocol": protocol,
        "port": port,
        "type": "ProtocolPortObject"
    }
    if Description != None:
        post_data['Description'] = Description
    post_url_data(name, api_path, post_data)

def create_PortGroup_object(name, post_data):
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/object/portobjectgroups'
    post_url_data(name, api_path, post_data)

def addACRule(Policies_name,sourcezonename,Source_name,destinationzonename,Destination_name,action, port_name=None,url_name=None):
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/policy/accesspolicies/' + Policy_UUID + '/accessrules'
    if 'group' in Source_name:
        SRC = get_Network_object_Group_ID(Source_name)
    else:
        SRC = get_Network_object_ID(Source_name)
    if 'group' in Destination_name:
        DES = get_Network_object_Group_ID(Destination_name)
    else:
        DES = get_Network_object_ID(Destination_name)
    sourcezoneUUID = get_zone_UUID(sourcezonename)
    destinationzoneUUID = get_zone_UUID(destinationzonename)
    if port_name != "None":
        post_data = {
            "sendEventsToFMC": True,
            "action": action,
            "enabled": True,
            "type": "AccessRule",
            "name": Policies_name,
            "logFiles": True,
            "logBegin": False,
            "logEnd": False,
            "sourceZones": {
                "objects": [
                    {
                        "name": sourcezonename,
                        "id": sourcezoneUUID,
                        "type": "SecurityZone"
                        }
                    ]
                },
            "destinationZones": {
                "objects": [
                    {
                        "name": destinationzonename,
                        "id": destinationzoneUUID,
                        "type": "SecurityZone"
                        }
                    ]
                },
            "variableSet": {
                "name": "Default Set",
                "id": "76fa83ea-c972-11e2-8be8-8e45bb1343c0",
                "type": "VariableSet"
                },
            "sourceNetworks": {
                "objects": [{
                    "type": SRC['type'],
                    "name": SRC['name'],
                    "id": SRC['id']
                    }]
                },
            "destinationNetworks": {
                "objects": [{
                    "type": DES['type'],
                    "name": DES['name'],
                    "id": DES['id']
                    }]
                },
            "destinationPorts": {
                "objects": []
                },
            }
        if 'group' in port_name:
            Des_Port = port_object_Group_ID(port_name)
            post_data["destinationPorts"]["objects"].append(Des_Port)
        else:
            Des_Port = port_object(port_name)
            post_data["destinationPorts"]["objects"].append(Des_Port)
    else:
        post_data = {
            "sendEventsToFMC": True,
            "action": action,
            "enabled": True,
            "type": "AccessRule",
            "name": Policies_name,
            "logFiles": True,
            "logBegin": False,
            "logEnd": False,
            "sourceZones": {
                "objects": [
                    {
                        "name": sourcezonename,
                        "id": sourcezoneUUID,
                        "type": "SecurityZone"
                        }
                    ]
                },
            "destinationZones": {
                "objects": [
                    {
                        "name": destinationzonename,
                        "id": destinationzoneUUID,
                        "type": "SecurityZone"
                        }
                    ]
                },
            "variableSet": {
                "name": "Default Set",
                "id": "76fa83ea-c972-11e2-8be8-8e45bb1343c0",
                "type": "VariableSet"
                },
            "sourceNetworks": {
                "objects": [{
                    "type": SRC['type'],
                    "name": SRC['name'],
                    "id": SRC['id']
                    }]
                },
            "destinationNetworks": {
                "objects": [{
                    "type": DES['type'],
                    "name": DES['name'],
                    "id": DES['id']
                    }]
                },
            "urls": {
                "objects": []
                },
            }
        if 'group' in url_name:
            url_para = get_urlGroup_object_ID(url_name)
            post_data["urls"]["objects"].append(url_para)
        else:
            url_para = get_url_object_ID(url_name)
            post_data["urls"]["objects"].append(url_para)
    post_url_data(Policies_name, api_path, post_data)

def create_urls_object(url_name, url_uri, Description = None):
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/object/urls'
    post_data = {
        "type": "Url",
        "overridable": False,
        "name": url_name,
        "url": url_uri
    }
    post_url_data(url_name, api_path, post_data)

def urls_group(url_group_name, post_data):
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/object/urlgroups'
    post_url_data(url_group_name, api_path, post_data)

def column_len2_fordef(i, var1, var2,vars):
    var1 = vars.cell(i, 1).value
    var2 = vars.cell(i, 2).value
    return [var1, var2]

def column_len3_fordef(i, var1, var2, var3 ,vars):
    var1 = vars.cell(i, 1).value
    var2 = vars.cell(i, 2).value
    var3 = vars.cell(i, 3).value
    return [var1, var2, var3]

def column_len4_fordef(i, var1, var2, var3, var4, vars):
    var1 = vars.cell(i, 1).value
    var2 = vars.cell(i, 2).value
    var3 = vars.cell(i, 3).value
    var4 = vars.cell(i, 4).value
    return [var1, var2, var3, var4]

Policy_UUID = []
file = r'C:\Users\Ein Lin\Desktop\python\FirePower\FMC_Policies_Scripts.xlsx'
wb = openpyxl.load_workbook(file)
device_info = wb.worksheets[0]
device_info_row = device_info.max_row

for i in range(2,device_info_row + 1):
    FMC_Addr = device_info.cell(row = i, column = 1).value
    username = device_info.cell(row = i, column = 2).value
    password = device_info.cell(row = i, column = 3).value

Server_url = 'https://' + FMC_Addr
r = requests.session()
fmc_gen_token(username,password)
headers = {'Content-Type': 'application/json', 'X-auth-access-token': FMC_token}
print('FMC current Version :',get_version())
get_current_Policy_UUID()

if len(Policy_UUID) == 1:
    Policy_UUID = Policy_UUID[0]
else:
    for i in Policy_UUID:
        print(i)
    Policy_UUID = input('which one Policy UUID you want to choice:\n')

#create network object , support network, range, host
print('\ncreate Network object\n')
ex_network = wb.worksheets[1]
ex_network_row = ex_network.max_row
for i in range(2, ex_network_row + 1):
    add_network_para = column_len3_fordef(i, 'network_name', 'network_ip', 'Description',ex_network)
    if add_network_para[2] == 'None':
        create_network_object(add_network_para[0], add_network_para[1])
    else:
        create_network_object(add_network_para[0], add_network_para[1], add_network_para[2])

#create network group objects
print('\ncreate Network Group objects \n')
ex_networkGroups = wb.worksheets[2]
ex_networkGroups_row = ex_networkGroups.max_row
ex_networkGroups_col = ex_networkGroups.max_column
for i in range(2, ex_networkGroups_row + 1):
    name = ex_networkGroups.cell(i, 1).value
    if name != 'None':
        post_data = {
            "name": name,
            "objects":[]
        }
        for j in range(2,ex_networkGroups_col + 1):
            object_IP_name = ex_networkGroups.cell(i, j).value
            if object_IP_name == None:
                break
            else:
                group_object = get_Network_object_ID(object_IP_name)
                group_object.pop('name')
                post_data["objects"].append(group_object)
        create_Networkgroups_object(name, post_data)

print('\ncreate Protocol object\n')
#create Protocol port
ex_Protocol = wb.worksheets[3]
ex_Protocol_row = ex_Protocol.max_row
for i in range(2, ex_Protocol_row + 1):
    add_port_para = column_len4_fordef(i, 'port_name','port_protocol','port','Description',ex_Protocol)
    if add_port_para[3] == 'None':
        create_port_object(add_port_para[0], add_port_para[1], add_port_para[2])
    else:
        create_port_object(add_port_para[0], add_port_para[1], add_port_para[2], add_port_para[3])

print('\ncreate Protocol Group\n)
ex_protocolGroups = wb.worksheets[4]
ex_protocolGroups_row = ex_protocolGroups.max_row
ex_protocolGroups_col = ex_protocolGroups.max_column
for i in range(2, ex_protocolGroups_row + 1):
    name = ex_protocolGroups.cell(i, 1).value
    if name != "None":
        post_data = {
            "name": name,
            "objects":[]
        }
        for j in range(2,ex_protocolGroups_col + 1):
            object_port_name = ex_protocolGroups.cell(i, j).value
            if object_port_name == None:
                break
            else:
                if object_port_name != 'ICMP':
                    group_port_object = port_object(object_port_name)
                    if group_port_object != None:
                        group_port_object.pop('name')
                        group_port_object.pop('protocol')
                    elif group_port_object == None:
                        group_port_object = research_port_object(object_port_name)
                    post_data["objects"].append(group_port_object)
                else:
                    group_port_object = get_ICMP_UUID()
                    post_data["objects"].append(group_port_object)
        create_PortGroup_object(name, post_data)

print('\ncreate url objects\n')
ex_urlObjects = wb.worksheets[5]
ex_urlObjects_row = ex_urlObjects.max_row
for i in range(2, ex_urlObjects_row + 1):
    add_url_para = column_len3_fordef(i, 'url_name','url_uri','Description',ex_urlObjects)
    print(add_url_para[0],add_url_para[1])
    if add_url_para[2] == 'None':
        create_urls_object(add_url_para[0], add_url_para[1])
    else:
        create_urls_object(add_url_para[0], add_url_para[1], add_url_para[2])

print("\nurl Groups\n")
ex_urlGroups = wb.worksheets[6]
ex_urlGroups_row = ex_urlGroups.max_row
ex_urlGroups_col = ex_urlGroups.max_column
for i in range(2, ex_urlGroups_row + 1):
    url_group_name = ex_urlGroups.cell(i, 1).value
    if url_group_name != "None":
        post_data = {
            "name": url_group_name,
            "objects": [],
            "type": "UrlGroup"
        }
        for j in range(2,ex_urlGroups_col + 1):
            object_url_name = ex_urlGroups.cell(i, j).value
            if object_url_name == None:
                break
            else:
                group_url_object = get_url_object_ID(object_url_name)
                group_url_object.pop('type')
                group_url_object.pop('name')
                post_data["objects"].append(group_url_object)
        urls_group(name, post_data)

print('\nadd AC Rules\n')
ex_add_AC_Rule = wb.worksheets[7]
ex_add_AC_Rule_row = ex_add_AC_Rule.max_row
time.sleep(3)
for i in range(2, ex_add_AC_Rule_row + 1):
    ex_Policies_name = ex_add_AC_Rule.cell(i, 1).value
    ex_sourcezonename = ex_add_AC_Rule.cell(i, 2).value
    ex_Source_name = ex_add_AC_Rule.cell(i, 3).value
    ex_destinationzonename = ex_add_AC_Rule.cell(i, 4).value
    ex_Destination_name = ex_add_AC_Rule.cell(i, 5).value
    ex_port_name = ex_add_AC_Rule.cell(i, 6).value
    ex_urlFiltering_name = ex_add_AC_Rule.cell(i, 7).value
    ex_action = ex_add_AC_Rule.cell(i, 8).value
    if ((ex_urlFiltering_name == "None") and (ex_port_name != "None")):
        addACRule(ex_Policies_name, ex_sourcezonename, ex_Source_name, ex_destinationzonename, ex_Destination_name, ex_action, ex_port_name)
    elif ((ex_urlFiltering_name != "None") and (ex_port_name == "None")):
        print(ex_port_name)
        addACRule(ex_Policies_name, ex_sourcezonename, ex_Source_name, ex_destinationzonename, ex_Destination_name, ex_action, ex_port_name, ex_urlFiltering_name)
    else:
        print("you have bad parameter")

print('Finished')
