import requests
import json
from requests.auth import HTTPBasicAuth
import time
import openpyxl

requests.packages.urllib3.disable_warnings()

def fmc_gen_token(username,passwd):
    global response, FMC_token, Domain_UUID, FMC_refresh_token
    token_api_path = "/api/fmc_platform/v1/auth/generatetoken"
    url = Server_url + token_api_path
    conn = r.post(url, auth=HTTPBasicAuth(username, passwd), verify=False)
    FMC_token = conn.headers["X-auth-access-token"]
    Domain_UUID = conn.headers["DOMAIN_UUID"]
    FMC_refresh_token = conn.headers["X-auth-refresh-token"]
    return FMC_token, Domain_UUID, FMC_refresh_token

def get_url_data(api_path):
    url = Server_url + api_path
    data = r.get(url, headers=headers, verify=False)
    json_data = json.loads(data.text)
    return json_data

def post_url_data(name, api_path, post_data):
    url = Server_url + api_path
    data = r.post(url, data=json.dumps(post_data), headers=headers, verify=False)
    if ((data.status_code == 200) or (data.status_code == 201)):
        print('post ' + name + 'Success, Fin')
    else:
        print('failed to Post')

def get_version():
    version_api = "/api/fmc_platform/v1/info/serverversion"
    version_json = get_url_data(version_api)
    FMC_Version = version_json['items'][0]['serverVersion']
    return FMC_Version

def get_current_Policy_UUID():
    get_Policy_UUID_url = '/api/fmc_config/v1/domain/' + Domain_UUID + '/policy/accesspolicies'
    Policy_UUID_json = get_url_data(get_Policy_UUID_url)
    for i in Policy_UUID_json['items']:
        print('name :', i['name'])
        print('ID :', i['id'])

def get_INT_UUID(Interface_Name):
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/devices/devicerecords/' + Device_UUID + '/physicalinterfaces'
    INT_UUID_json = get_url_data(api_path)
    try:
        if INT_UUID_json['items'] != []:
            for i in INT_UUID_json['items']:
                if i['name'] == Interface_Name:
                    Interface_UUID = i['id']
                    return Interface_UUID
    except:
        return False


def get_Network_object_ID(net_name) -> dict:
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID +'/object/networks'
    get_ID_json = get_url_data(api_path)
    for i in get_ID_json['items']:
        if net_name == i['name']:
            return { "type": i['type'],
                     "name": i['name'],
                     "id": i['id']}

def get_zone_UUID(zone_name):
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/object/securityzones'
    get_zone_json = get_url_data(api_path)
    for i in get_zone_json['items']:
        if zone_name == i['name']:
            return i['id']

def port_object(port_name):
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/object/protocolportobjects'
    port_info_json = get_url_data(api_path)
    for i in port_info_json['items']:
        if port_name == i['name']:
            protocol_url = i['links']['self']
            port_info_d = r.get(url, headers = headers, verify = False)
            json_port_info_d = json.loads(json_port_info_d)
            return {
                'type':json_port_info_d['type'],
                'protocol':json_port_info_d['protocol'],
                'name':json_port_info_d['name'],
                'id':json_port_info_d['id']
            }

def add_Device(Device_name,host_name,reg_Key,natID=None):
    #global Device_UUID
    api_path = '/api/fmc_config/v1/domain/e276abec-e0f2-11e3-8169-6d9ed49b625f/devices/devicerecords'
    url = Server_url + api_path
    if natID == None:
        post_data = {
            "name": Device_name,
            "hostName": host_name,
            "regKey": reg_Key,
            "type": "Device",
            "license_caps": [
                "BASE",
                "MALWARE",
                "URLFilter",
                "THREAT"
                ],
            "accessPolicy": {
                "id": policy_UUID,
                "type": "AccessPolicy"
                }
            }
    else:
        post_data['natID'] = natID

    Device_Add = r.post(url, data=json.dumps(post_data), headers=headers, verify=False)
    '''
    try:
        Device_Add
    except:
        conn_failed = True
        while conn_failed:
            Device_Add = r.post(url, data=json.dumps(post_data), headers=headers, verify=False)
            if ((Device_Add.status_code == 201) or (Device_Add.status_code == 202)):
                conn_failed = False
            else:
                conn_failed = True
    '''
    if ((Device_Add.status_code == 201) or (Device_Add.status_code == 202)):
        process_not_finished = True
        while process_not_finished:
            process_reg = r.get(url, headers=headers, verify=False)
            json_process_reg = json.loads(process_reg.text)
            try:
                for i in json_process_reg['items']:
                    if i['name'] == post_data['name']:
                        process_not_finished = False
                        Device_UUID = i['id']
                        print('success add Device')
                        return Device_UUID
            except:
                print('waiting for add Device Process')
                process_not_finished = True

def create_Policy(policy_name):
    policy_api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/policy/accesspolicies'
    Policy_url = Server_url + policy_api_path
    post_data = {
        "type": "AccessPolicy",
        "name": policy_name,
        "description": "Enterprise Policy to Detect and Prevent Threats",
        "defaultAction": {
            "intrusionPolicy": {
                "name": "Security Over Connectivity",
                "id": "abba9b63-bb10-4729-b901-2e2aa0f4491c",
                "type": "IntrusionPolicy"
            },
            "variableSet": {
                "name": "Default Set",
                "id": "76fa83ea-c972-11e2-8be8-8e45bb1343c0",
                "type": "VariableSet"
            },
            "type": "AccessPolicyDefaultAction",
            "logBegin": False,
            "logEnd": True,
            "sendEventsToFMC": True
        }
    }
    create_r = r.post(Policy_url, data=json.dumps(post_data), headers=headers, verify=False)
    if ((create_r.status_code == 201) or (create_r.status_code == 202)):
        print('Success Create Policy')
        json_r = json.loads(create_r.text)
        policy_UUID = json_r['id']
    else:
        print('Failed to Create Policy')
    return policy_UUID

def create_network_object(name,ip_address, Description = None):
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/object/networks'
    post_data = {
        "name": name,
        "value": ip_address,
        "overridable": False,
        "description": Description,
        "type": "Network"
        }
    post_url_data(ip_address, api_path, post_data)
    #json_Create_Network_object = json.loads(Create_Network_object.text)

def addACRule(Policies_name,sourcezonename,Source_name,destinationzonename,Destination_name,port_name,action):
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/policy/accesspolicies/' + policy_UUID + '/accessrules'
    SRC = get_Network_object_ID(Source_name)
    DES = get_Network_object_ID(Destination_name)
    Des_Port = port_object(port_name)
    sourcezoneUUID = get_zone_UUID(sourcezonename)
    destinationzoneUUID = get_zone_UUID(destinationzonename)
    post_data = {
        "sendEventsToFMC": True,
        "action": action,
        "enabled": True,
        "type": "AccessRule",
        "name": Policies_name,
        "logFiles": True,
        "logBegin": False,
        "logEnd": False,
        "sourceZones": {
            "objects": [
                {
                    "name": sourcezonename,
                    "id": sourcezoneUUID,
                    "type": "SecurityZone"
                }
            ]
        },
        "destinationZones": {
            "objects": [
                {
                    "name": destinationzonename,
                    "id": destinationzoneUUID,
                    "type": "SecurityZone"
                }
            ]
        },
        "variableSet": {
            "name": "Default Set",
            "id": "76fa83ea-c972-11e2-8be8-8e45bb1343c0",
            "type": "VariableSet"
        },
        "sourceNetworks": {
            "objects": [{
                "type": SRC['type'],
                "name": SRC['name'],
                "id": SRC['id']
            }]
        },
        "destinationNetworks": {
            "objects": [{
                "type": DES['type'],
                "name": DES['name'],
                "id": DES['id']
            }]
        },
        "destinationPorts": {
            "objects": [
                {
                "type": Des_Port['type'],
                "protocol": Des_Port['protocol'],
                "name": Des_Port['name'],
                "id": Des_Port['id']
                }]
            },
        }
    post_url_data(Policies_name, api_path, post_data)
    #add_ACL = r.post(url, data=json.dumps(post_data), headers=headers, verify=False)
    #if ((add_ACL.status_code == 200) or (add_ACL.status_code == 201)):
    #    print('Add_success')
    #else:
    #    print(add_ACL.status_code,'\n')
    #    print(add_ACL.text)

def Create_Zones(zone_name, interfaceMode='ROUTED'):
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/object/securityzones'
    post_data = {
        "type": "SecurityZone",
        "name": zone_name,
        "interfaceMode": interfaceMode
    }
    post_url_data(zone_name,api_path, post_data)

def assign_zones(ifname, Interface_Name, Zone_Name):
    Interface_UUID = False
    while Interface_UUID == False:
        Interface_UUID = get_INT_UUID(Interface_Name)
        if Interface_UUID == False:
            print('waiting for FMC process')
        time.sleep(2)
    Interface_UUID = get_INT_UUID(Interface_Name)
    Zone_UUID = get_zone_UUID(Zone_Name)
    api_path = '/api/fmc_config/v1/domain/' + Domain_UUID + '/devices/devicerecords/' + Device_UUID + '/physicalinterfaces/' + Interface_UUID
    url = Server_url + api_path
    post_data = {
        "mode": "NONE",
        "enabled": True,
        "MTU": 1500,
        "ifname": ifname,
        "fragmentReassembly": False,
        "enableDNSLookup": False,
        "enableAntiSpoofing": False,
        "managementOnly": False,
        "securityZone": {
            "id": Zone_UUID,
            "type": "SecurityZone"
            },
            "name": Interface_Name,
            "id": Interface_UUID
            }
    int_Zones = r.put(url, data=json.dumps(post_data), headers=headers, verify=False)
    if ((int_Zones.status_code >= 200) or (int_Zones.status_code <= 299)):
        print('Put ' + Interface_Name + ' into ' + Zone_Name + ' successful ')
    else:
        print(int_Zones.text)

def column_len4_fordef(i,var1,var2,var3,var4,vars):
    var1 = vars.cell(i,  1).value
    var2 = vars.cell(i,  2).value
    var3 = vars.cell(i,  3).value
    var4 = vars.cell(i,  4).value
    return [var1,var2,var3,var4]

def column_len2_fordef(i,var1,var2,vars):
    var1 = vars.cell(i,  1).value
    var2 = vars.cell(i,  2).value
    return [var1,var2]

file = r'C:\Users\Ein Lin\Desktop\python\FirePower\Setup_INFO.xlsx'
wb = openpyxl.load_workbook(file)
ex_device_info = wb.worksheets[0]
ex_add_devices = wb.worksheets[1]
ex_zones_name = wb.worksheets[2]
ex_assign_zone = wb.worksheets[3]
ex_add_AC_Rule = wb.worksheets[4]
ex_device_info_row = ex_device_info.max_row
ex_add_devices_row = ex_add_devices.max_row
ex_zones_name_row = ex_zones_name.max_row
ex_assign_zone_row = ex_assign_zone.max_row
ex_add_AC_Rule_row = ex_add_AC_Rule.max_row

for i in range(2,ex_device_info_row + 1):
    ex_device_info_list = column_len4_fordef(i,'FMC_Addr','username','password','policy_name',ex_device_info)

Server_url = 'https://' + ex_device_info_list[0]
r = requests.session()
fmc_gen_token(ex_device_info_list[1],ex_device_info_list[2])
headers = {'Content-Type': 'application/json', 'X-auth-access-token': FMC_token}
print('FMC current Version :',get_version())
policy_UUID = create_Policy(ex_device_info_list[3])
time.sleep(5)

for i in range(2,ex_add_devices_row + 1):
    add_device_para = column_len4_fordef(i, 'Device_Name','host_name','reg_Key','natID',ex_add_devices)
    #add Device , you need type in command 'config manager add FMC_add regKey' in FTD first
    if add_device_para[3] == 'None':
        Device_UUID = add_Device(add_device_para[0],add_device_para[1],add_device_para[2])
    else:
        Device_UUID = add_Device(add_device_para[0], add_device_para[1], add_device_para[2], add_device_para[3])

time.sleep(3)
for i in range(2, ex_zones_name_row +1 ):
    zone_name_para = column_len2_fordef(i, 'zone_name', 'interfaceMode', ex_zones_name)
    Create_Zones(zone_name_para[0],zone_name_para[1])
'''
for i in range(2, assign_zone_row +1 ):
    assign_zone_para = column_len2_fordef(i, 'interface_name', 'zone_name',assign_zone)
    assign_zones(assign_zone_para[0], assign_zone_para[1])
    time.sleep(1)
'''
for i in range(2, ex_assign_zone_row + 1):
    ex_Port_ID = ex_assign_zone.cell(i, 1).value
    ex_zone_name = ex_assign_zone.cell(i, 2).value
    ex_if_name = ex_assign_zone.cell(i, 3).value
    assign_zones(ex_if_name, ex_Port_ID, ex_zone_name)
    time.sleep(1)

time.sleep(3)
for i in range(2, ex_add_AC_Rule_row + 1):
    ex_Policies_name = ex_add_AC_Rule.cell(i, 1).value
    ex_sourcezonename = ex_add_AC_Rule.cell(i, 2).value
    ex_Source_name = ex_add_AC_Rule.cell(i, 3).value
    ex_destinationzonename = ex_add_AC_Rule.cell(i, 4).value
    ex_Destination_name = ex_add_AC_Rule.cell(i, 5).value
    ex_port_name = ex_add_AC_Rule.cell(i, 6).value
    ex_action = ex_add_AC_Rule.cell(i, 7).value
    addACRule(ex_Policies_name, ex_sourcezonename, ex_Source_name, ex_destinationzonename, ex_Destination_name, ex_port_name, ex_action)
